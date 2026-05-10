"""
Read both slip trackers and emit static HTML viewer pages + a stats JSON for
the portal home page. Idempotent — runs as part of the Friday autonomous flow.

Output:
  flomatic-reporting/petrol-slips.html
  flomatic-reporting/other-slips.html
  flomatic-reporting/data/portal-stats.json
"""
from pathlib import Path
from datetime import datetime
from html import escape
import json
from openpyxl import load_workbook

WORKSPACE = Path(r"C:\Users\quint\OneDrive\2.Areas\Flomatic")
REPO      = Path(r"C:\Users\quint\flomatic-reporting")
DATA_DIR  = REPO / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

PETROL_XLSX = WORKSPACE / "Reports" / "petrol_slip_tracking.xlsx"
OTHER_XLSX  = WORKSPACE / "Reports" / "other_slips_tracking.xlsx"

# Column slices: data rows start at 5 in petrol (header rows 1-4 incl. description row);
# at 4 in the new other tracker.
PETROL_HEADERS = [
    "Slip ID","Date","Time","Station","Address","Town/City","Brand","Grade",
    "Litres","R/L","Total (R)","Pay Method","Card","Tx Ref","Vehicle Reg",
    "Odometer","Driver","Business Purpose","Zoho Account","Linked Bank Txn",
    "Slip File","Status","Notes",
]
OTHER_HEADERS = [
    "Slip ID","Date","Time","Vendor","Address","Town/City","Vendor Type",
    "Total (R)","VAT (R)","Pay Method","Card","Tx Ref","Tax Invoice",
    "Driver","Business Purpose","Zoho Account","Linked Bank Txn",
    "Slip File","Status","Notes",
]

def read_rows(path: Path, sheet_name: str | None, data_start_row: int, ncols: int):
    """Return list of lists. Skip empty rows (no Slip ID)."""
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if not row or not row[0]:
            continue
        # Right-pad / trim to ncols
        vals = list(row[:ncols]) + [None] * max(0, ncols - len(row))
        rows.append(vals)
    return rows

def fmt_cell(val):
    if val is None:
        return ""
    if isinstance(val, (datetime,)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        if val.is_integer():
            return f"{int(val)}"
        return f"{val:,.2f}"
    return str(val)

def status_class(status: str) -> str:
    s = (status or "").strip().lower()
    if s in ("pending",): return "status-pending"
    if s in ("hold",): return "status-hold"
    if s in ("needs review", "needs-review"): return "status-needs-review"
    if s in ("rejected",): return "status-rejected"
    if s in ("reviewed", "approved"): return "status-reviewed"
    if s in ("imported",): return "status-imported"
    return ""

def row_class(status: str) -> str:
    s = (status or "").strip().lower()
    if s == "rejected": return "row-rejected"
    if s in ("needs review", "needs-review", "hold", "pending"): return "row-needs-review"
    return ""

def render_table(headers, rows, status_idx):
    out = ['<table class="viewer-table"><thead><tr>']
    out += [f"<th>{escape(h)}</th>" for h in headers]
    out += ["</tr></thead><tbody>"]
    if not rows:
        out += [f'<tr><td colspan="{len(headers)}" style="text-align:center;color:var(--muted);padding:32px;">No rows yet — Friday autonomous run will populate this.</td></tr>']
    for row in rows:
        status = row[status_idx] if status_idx < len(row) else ""
        rc = row_class(status)
        out += [f'<tr class="{rc}">']
        for i, v in enumerate(row):
            text = escape(fmt_cell(v))
            if i == status_idx:
                cls = status_class(v)
                out += [f'<td class="{cls}">{text}</td>']
            else:
                out += [f"<td>{text}</td>"]
        out += ["</tr>"]
    out += ["</tbody></table>"]
    return "".join(out)

def page_html(title, eyebrow, blurb, headers, rows, status_idx, total_idx):
    total_amount = 0.0
    pending_count = 0
    needs_review_count = 0
    for r in rows:
        try:
            total_amount += float(r[total_idx]) if r[total_idx] not in (None, "") else 0.0
        except (TypeError, ValueError):
            pass
        s = (r[status_idx] or "").strip().lower() if status_idx < len(r) else ""
        if s in ("pending", "hold"):
            pending_count += 1
        if s in ("needs review", "needs-review"):
            needs_review_count += 1

    table = render_table(headers, rows, status_idx)
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(title)} — Forflowmatic</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Space+Grotesk:wght@400;500;600;700&family=DM+Sans:wght@300;400;500;600&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
  <link rel="stylesheet" href="assets/style.css">
</head>
<body>
<header class="flo-header">
  <div class="flo-wordmark">
    <span class="flo-name">FORFLOMATIC</span>
    <span class="flo-tagline">Reporting</span>
  </div>
  <div class="flo-status-pills">
    <span class="flo-pill success">Live</span>
  </div>
</header>
<main>
  <a class="viewer-back" href="index.html">← Back to portal</a>
  <p class="label-eyebrow">{escape(eyebrow)}</p>
  <h1 class="flo-h1">{escape(title)}</h1>
  <p class="flo-sub">{escape(blurb)}</p>

  <div class="viewer-stat-row">
    <div class="viewer-stat"><span class="label">Total slips</span><span class="value">{len(rows)}</span></div>
    <div class="viewer-stat"><span class="label">Pending / hold</span><span class="value" style="color:var(--flo-amber-alert);">{pending_count}</span></div>
    <div class="viewer-stat"><span class="label">Needs review</span><span class="value" style="color:var(--flo-fault-red);">{needs_review_count}</span></div>
    <div class="viewer-stat"><span class="label">Total amount</span><span class="value">R {total_amount:,.2f}</span></div>
  </div>

  <div class="viewer-toolbar">
    <input id="search" type="text" class="viewer-search" placeholder="Search by any field…">
  </div>

  {table}
</main>
<footer class="flo-footer">
  Forflowmatic Automations · Reporting Portal · Auto-generated <span id="genTs"></span>
</footer>
<script>
  document.getElementById('genTs').textContent = '{datetime.now().strftime("%Y-%m-%d %H:%M")}';
  const search = document.getElementById('search');
  const rows = document.querySelectorAll('.viewer-table tbody tr');
  search.addEventListener('input', () => {{
    const q = search.value.trim().toLowerCase();
    rows.forEach(r => {{
      r.style.display = (!q || r.textContent.toLowerCase().includes(q)) ? '' : 'none';
    }});
  }});
</script>
</body>
</html>
"""

def main():
    petrol_rows = read_rows(PETROL_XLSX, "Petrol Slip Tracker", 5, len(PETROL_HEADERS))
    other_rows  = read_rows(OTHER_XLSX,  "Other Slip Tracker",  4, len(OTHER_HEADERS))

    # Petrol: status idx = 21, total idx = 10
    (REPO / "petrol-slips.html").write_text(
        page_html(
            "Petrol Slip Tracker",
            "Fuel · Travel",
            "Every fuel-station slip captured by the Friday autonomous run. Pre-import gate held — Zoho status per row.",
            PETROL_HEADERS, petrol_rows,
            status_idx=21, total_idx=10,
        ),
        encoding="utf-8",
    )

    # Other: status idx = 18, total idx = 7
    (REPO / "other-slips.html").write_text(
        page_html(
            "Other Expense Slips",
            "Expense · Other",
            "Non-petrol receipts (food, groceries, pharmacy, hardware, optometry, etc.). Routed here automatically when the Friday run can't classify a slip as fuel.",
            OTHER_HEADERS, other_rows,
            status_idx=18, total_idx=7,
        ),
        encoding="utf-8",
    )

    # Stats JSON for the portal home cards
    def count_status(rows, idx, *targets):
        return sum(1 for r in rows
                   if (r[idx] or "").strip().lower() in {t.lower() for t in targets})

    stats = {
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "petrol": {
            "total":   len(petrol_rows),
            "pending": count_status(petrol_rows, 21, "Pending", "Hold"),
        },
        "other": {
            "total":        len(other_rows),
            "needs_review": count_status(other_rows, 18, "Needs review", "needs-review", "Hold"),
        },
    }
    (DATA_DIR / "portal-stats.json").write_text(json.dumps(stats, indent=2), encoding="utf-8")

    print(f"[ok] petrol-slips.html: {len(petrol_rows)} rows")
    print(f"[ok] other-slips.html:  {len(other_rows)} rows")
    print(f"[ok] data/portal-stats.json")
    print(f"[ok] generated at {stats['updated']}")

if __name__ == "__main__":
    main()
